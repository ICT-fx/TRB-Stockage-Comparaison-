"""
TRB Chemedica — Stock Comparison API
Compare theoretical (Proconcept) vs actual (RK Logistik) inventory.
Comparison key: SKU + Lot number.
"""

import io
import re
import zipfile
from typing import Any

import pandas as pd
from fastapi import FastAPI, File, Form, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

import templates as template_store
import comments as comment_store

# Patch zipfile to ignore CRC-32 errors commonly found in ERP-exported Excel files
zipfile.ZipExtFile._update_crc = lambda *args, **kwargs: None

app = FastAPI(title="TRB Stock Compare API")

# CORS restreint : seules les origines connues peuvent appeler l'API.
# - le frontend hébergé (Render, sous-domaine *.onrender.com du projet) ;
# - le poste local (l'.exe sert l'interface sur localhost:8000 ; l'ancien mode
#   dev utilisait localhost:3000).
# Les routes /templates modifient un état persistant : on bloque donc les sites
# tiers. L'app n'utilise ni cookie ni session -> allow_credentials=False.
app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=(
        r"^https://trb-stock-compare-front[a-z0-9-]*\.onrender\.com$"
        r"|^http://(localhost|127\.0\.0\.1):(8000|3000)$"
    ),
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def _is_numeric_code(val: Any) -> bool:
    """Return True if *val* looks like a numeric product code (e.g. '1349')."""
    if val is None:
        return False
    s = str(val).strip().replace(".0", "")
    return bool(re.fullmatch(r"\d+", s))


def _clean_code(val: Any) -> str:
    """Convert a code value to a clean integer string, e.g. 687.0 → '687'."""
    return str(int(float(str(val).strip())))


def _clean_lot(val: Any) -> str:
    """Normalize a lot number to a clean string. Handles floats like 462994.0."""
    if pd.isna(val):
        return ""
    s = str(val).strip()
    # If it looks like a float integer (e.g. '412561.0'), remove the .0
    if re.fullmatch(r"\d+\.0", s):
        return s[:-2]
    return s


def _format_date_yyyymmdd(val: Any) -> str:
    """Convert a YYYYMMDD integer (e.g. 20280731) to DD.MM.YYYY string."""
    try:
        s = str(int(float(str(val))))
        if len(s) == 8:
            return f"{s[6:8]}.{s[4:6]}.{s[0:4]}"
    except Exception:
        pass
    return str(val)


def _parse_rk_date(val: Any) -> str:
    """Parse RK Logistik date formats to DD.MM.YYYY.
    Handles: '2028-03' (YYYY-MM), '2028-01-12 00:00:00', pandas Timestamp.
    """
    if pd.isna(val):
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%d.%m.%Y")
    s = str(val).strip()
    if re.fullmatch(r"\d{4}-\d{2}", s):
        return f"01.{s[5:7]}.{s[0:4]}"
    try:
        dt = pd.to_datetime(s, dayfirst=True)
        return dt.strftime("%d.%m.%Y")
    except Exception:
        return s


# ──────────────────────────────────────────────
# New parsers — Lot-number based
# ──────────────────────────────────────────────

def parse_proconcept_lot(raw_bytes: bytes) -> tuple[list[dict], pd.DataFrame]:
    """
    Parse the new Proconcept layout with lot numbers.
    Columns (header=0):
      [0] Stock, [1] Emplacement, [2] Description courte, [3] Chronologie (YYYYMMDD),
      [4] Référence principale (SKU), [5] Numéro de lot, [6] Version, [7] Qté effective

    Lot number = col[5] (Numéro de lot) if not NaN, else col[6] (Version).
    Returns (list_of_dicts, raw_dataframe).
    """
    df = pd.read_excel(io.BytesIO(raw_bytes), header=0)

    if len(df.columns) < 8:
        raise ValueError(
            f"Fichier Proconcept (lots) invalide : {len(df.columns)} colonnes détectées, "
            "8 attendues (Stock, Emplacement, Description, Chronologie, SKU, N°Lot, Version, Qté)."
        )

    col_code = df.columns[4]   # Référence principale
    col_lot1 = df.columns[5]   # Numéro de lot
    col_lot2 = df.columns[6]   # Version
    col_qty  = df.columns[7]   # Qté effective
    col_desc = df.columns[2]   # Description courte
    col_date = df.columns[3]   # Chronologie

    # Forward-fill description
    df[col_desc] = df[col_desc].ffill()

    products: list[dict] = []

    for _, row in df.iterrows():
        code_raw = row[col_code]
        if not _is_numeric_code(code_raw):
            continue

        # Lot = Numéro de lot si non vide, sinon Version
        lot1 = _clean_lot(row[col_lot1])
        lot2 = _clean_lot(row[col_lot2])
        lot = lot1 if lot1 else lot2
        if not lot:
            continue  # Skip rows without a lot number

        code = _clean_code(code_raw)
        qty  = int(row[col_qty]) if pd.notna(row[col_qty]) else 0
        desc = str(row[col_desc]).strip() if pd.notna(row[col_desc]) else ""
        date_str = _format_date_yyyymmdd(row[col_date]) if pd.notna(row[col_date]) else ""

        products.append({
            "code": code,
            "lot": lot,
            "date": date_str,
            "qty": qty,
            "description": desc,
        })

    return products, df


def parse_rk_lot(raw_bytes: bytes) -> tuple[list[dict], pd.DataFrame]:
    """
    Parse the new RK Logistik layout with lot numbers.
    The real header is on row index 1 (row 0 is empty), data starts at row 2.
    Columns after skipping:
      [0] SKU, [1] Lot (Lagerort), [2] Date (G/YYYY-MM), [3] Description (Kurztext),
      [4] Quantity (Bestand), [5] Unit (ignored), [6] Lot/Exp combined (ignored)

    Returns (list_of_dicts, raw_dataframe_with_real_header).
    """
    df_raw = pd.read_excel(io.BytesIO(raw_bytes), header=None)

    # Find the actual header row: first row where col[0] is NaN and col[1] == 'Lagerort'
    # Typically row index 1. We skip row 0 (empty) and use row 1 as header.
    df = pd.read_excel(io.BytesIO(raw_bytes), header=1)

    if len(df.columns) < 5:
        raise ValueError(
            f"Fichier RK Logistik (lots) invalide : {len(df.columns)} colonnes détectées, "
            "au moins 5 attendues (SKU, Lot, Date, Description, Quantité)."
        )

    col_code = df.columns[0]   # SKU
    col_lot  = df.columns[1]   # Lagerort = Lot number
    col_date = df.columns[2]   # G = Date YYYY-MM
    col_desc = df.columns[3]   # Kurztext = Description
    col_qty  = df.columns[4]   # Bestand = Quantity

    products: list[dict] = []

    for _, row in df.iterrows():
        code_raw = row[col_code]
        if not _is_numeric_code(code_raw):
            continue

        lot = _clean_lot(row[col_lot])
        if not lot:
            continue

        code = _clean_code(code_raw)
        qty  = int(row[col_qty]) if pd.notna(row[col_qty]) else 0
        desc = str(row[col_desc]).strip() if pd.notna(row[col_desc]) else ""
        date_str = _parse_rk_date(row[col_date])

        products.append({
            "code": code,
            "lot": lot,
            "date": date_str,
            "qty": qty,
            "description": desc,
        })

    return products, df


def parse_storage_with_template(raw_bytes: bytes, template: dict) -> tuple[list[dict], pd.DataFrame]:
    """
    Parse un fichier d'espace de stockage selon un template de colonnes.
    template = {"header_row": int(1-based), "columns": {sku,lot,qty,date,description}}
    (date/description peuvent être None). Renvoie la même structure que parse_rk_lot.
    """
    cols = template["columns"]
    header_row = int(template.get("header_row", 1))
    df = pd.read_excel(io.BytesIO(raw_bytes), header=header_row - 1)

    used = [cols["sku"], cols["lot"], cols["qty"]]
    used += [cols[k] for k in ("date", "description") if cols.get(k) is not None]
    max_idx = max(used)
    if len(df.columns) <= max_idx:
        raise ValueError(
            f"Le fichier n'a que {len(df.columns)} colonnes, "
            f"le template en attend au moins {max_idx + 1}."
        )

    products: list[dict] = []
    for _, row in df.iterrows():
        code_raw = row.iloc[cols["sku"]]
        if not _is_numeric_code(code_raw):
            continue
        lot = _clean_lot(row.iloc[cols["lot"]])
        if not lot:
            continue
        code = _clean_code(code_raw)
        qty_cell = row.iloc[cols["qty"]]
        qty = int(qty_cell) if pd.notna(qty_cell) else 0

        date_str = ""
        if cols.get("date") is not None:
            date_str = _parse_rk_date(row.iloc[cols["date"]])

        desc = ""
        if cols.get("description") is not None:
            dv = row.iloc[cols["description"]]
            desc = str(dv).strip() if pd.notna(dv) else ""

        products.append({
            "code": code, "lot": lot, "date": date_str,
            "qty": qty, "description": desc,
        })

    return products, df


# ──────────────────────────────────────────────
# Comparison logic — SKU + Lot number
# ──────────────────────────────────────────────

def compare_by_lot(
    theoretical: list[dict],
    actual: list[dict],
) -> dict:
    """
    Compare two stock lists by SKU + lot number.
    Key = (code, lot). Exact match required.
    """
    def _build_map(items: list[dict]) -> dict[tuple, dict]:
        m: dict[tuple, dict] = {}
        for item in items:
            key = (item["code"], item["lot"])
            if key in m:
                m[key]["qty"] += item["qty"]
            else:
                m[key] = {**item}
        return m

    theo_map = _build_map(theoretical)
    actual_map = _build_map(actual)

    ok = []
    discrepancies = []

    all_keys = sorted(set(theo_map) | set(actual_map))

    for key in all_keys:
        code, lot = key
        in_theo  = key in theo_map
        in_actual = key in actual_map

        if in_theo and in_actual:
            t = theo_map[key]
            a = actual_map[key]
            qty_theo = t["qty"]
            qty_real = a["qty"]
            delta = qty_real - qty_theo
            entry = {
                "code": code,
                "lot": lot,
                "date_proconcept": t.get("date", ""),
                "date_rk": a.get("date", ""),
                "description_theorique": t.get("description", ""),
                "description_reel": a.get("description", ""),
                "qty_theorique": qty_theo,
                "qty_reel": qty_real,
                "delta": delta,
            }
            if delta == 0:
                ok.append(entry)
            else:
                discrepancies.append(entry)

        elif in_theo:
            t = theo_map[key]
            discrepancies.append({
                "code": code,
                "lot": lot,
                "date_proconcept": t.get("date", ""),
                "date_rk": "",
                "description_theorique": t.get("description", ""),
                "description_reel": "",
                "qty_theorique": t["qty"],
                "qty_reel": 0,
                "delta": -t["qty"],
            })
        else:
            a = actual_map[key]
            discrepancies.append({
                "code": code,
                "lot": lot,
                "date_proconcept": "",
                "date_rk": a.get("date", ""),
                "description_theorique": "",
                "description_reel": a.get("description", ""),
                "qty_theorique": 0,
                "qty_reel": a["qty"],
                "delta": a["qty"],
            })

    total = len(ok) + len(discrepancies)
    stats = {
        "total_products": total,
        "ok_count": len(ok),
        "discrepancy_count": len(discrepancies),
        "match_rate": round(len(ok) / total * 100, 1) if total else 0,
    }

    return {
        "has_lots": True,
        "ok": ok,
        "discrepancies": discrepancies,
        "stats": stats,
    }


# ──────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────

_HEADER_FONT      = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_TRB_BLUE   = "004B87"
_GREEN      = "2E7D32"
_ORANGE     = "F57F17"
_GREY       = "607D8B"


def _style_header(ws, fill_color: str, col_count: int):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = fill
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 55)


def _write_raw_sheet(wb: Workbook, title: str, df: pd.DataFrame, fill_color: str):
    """Write a raw dataframe (as-is) to a new sheet with styled header."""
    ws = wb.create_sheet(title)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    _style_header(ws, fill_color, len(df.columns))
    _auto_width(ws)


def build_excel(result: dict, df_proconcept: pd.DataFrame, df_rk: pd.DataFrame) -> bytes:
    """Build a styled .xlsx workbook from comparison results + raw source sheets."""
    wb = Workbook()

    headers_base = ["Code", "N° de lot", "Date Proconcept", "Date RK", "Description", "Qté Proconcept", "Qté Réelle", "Delta"]

    # ── OK ──
    ws_ok = wb.active
    ws_ok.title = "OK"
    ws_ok.append(headers_base)
    for item in result["ok"]:
        ws_ok.append([
            item["code"], item["lot"],
            item.get("date_proconcept", ""), item.get("date_rk", ""),
            item.get("description_theorique") or item.get("description_reel", ""),
            item["qty_theorique"], item["qty_reel"], item["delta"],
        ])
    _style_header(ws_ok, _GREEN, len(headers_base))
    _auto_width(ws_ok)

    # ── Écarts ──
    ws_disc = wb.create_sheet("Écarts")
    ws_disc.append(headers_base)
    for item in result["discrepancies"]:
        ws_disc.append([
            item["code"], item["lot"],
            item.get("date_proconcept", ""), item.get("date_rk", ""),
            item.get("description_theorique") or item.get("description_reel", ""),
            item["qty_theorique"], item["qty_reel"], item["delta"],
        ])
    _style_header(ws_disc, _ORANGE, len(headers_base))
    _auto_width(ws_disc)

    # ── Données brutes Proconcept ──
    _write_raw_sheet(wb, "Données Proconcept", df_proconcept, _TRB_BLUE)

    # ── Données brutes RK ──
    _write_raw_sheet(wb, "Données RK Logistik", df_rk, _GREY)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────
# Comparison runner
# ──────────────────────────────────────────────

def _run_comparison(theo_bytes: bytes, real_bytes: bytes, storage_template: dict) -> tuple[dict, pd.DataFrame, pd.DataFrame]:
    """Parse both files and run lot-based comparison. Returns (result, df_pro, df_rk)."""
    theo_list, df_pro = parse_proconcept_lot(theo_bytes)
    actual_list, df_rk = parse_storage_with_template(real_bytes, storage_template)
    result = compare_by_lot(theo_list, actual_list)
    return result, df_pro, df_rk


# ──────────────────────────────────────────────
# API Routes
# ──────────────────────────────────────────────

@app.post("/compare")
async def compare(
    file_theorique: UploadFile = File(...),
    file_reel: UploadFile = File(...),
    storage_template_id: str = Form("basic-stock"),
):
    """Compare two Excel files and return JSON results."""
    tpl = template_store.get_template(storage_template_id)
    if tpl is None:
        raise HTTPException(status_code=400, detail=f"Template introuvable : {storage_template_id}")
    try:
        theo_bytes = await file_theorique.read()
        real_bytes = await file_reel.read()
        result, _, _ = _run_comparison(theo_bytes, real_bytes, tpl)
        for d in result["discrepancies"]:
            d["stored_comment"] = comment_store.get_comment(d["code"], d["lot"])
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lors du traitement : {str(e)}")


@app.post("/compare/download")
async def compare_download(
    file_theorique: UploadFile = File(...),
    file_reel: UploadFile = File(...),
    storage_template_id: str = Form("basic-stock"),
):
    """Compare two Excel files and return an Excel report with raw data sheets."""
    tpl = template_store.get_template(storage_template_id)
    if tpl is None:
        raise HTTPException(status_code=400, detail=f"Template introuvable : {storage_template_id}")
    try:
        theo_bytes = await file_theorique.read()
        real_bytes = await file_reel.read()
        result, df_pro, df_rk = _run_comparison(theo_bytes, real_bytes, tpl)
        excel_bytes = build_excel(result, df_pro, df_rk)

        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=comparaison_stock.xlsx"},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lors du traitement : {str(e)}")



# ──────────────────────────────────────────────
# Templates d'espace de stockage
# ──────────────────────────────────────────────

@app.get("/templates")
async def list_templates():
    return {"templates": template_store.all_templates()}


@app.post("/templates")
async def create_template_route(payload: dict):
    try:
        return template_store.create_template(payload)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.put("/templates/{template_id}")
async def update_template_route(template_id: str, payload: dict):
    try:
        return template_store.update_template(template_id, payload)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except KeyError:
        raise HTTPException(status_code=404, detail="Template introuvable.")


@app.delete("/templates/{template_id}")
async def delete_template_route(template_id: str):
    try:
        template_store.delete_template(template_id)
        return {"ok": True}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except KeyError:
        raise HTTPException(status_code=404, detail="Template introuvable.")


@app.post("/templates/preview")
async def preview_template(file: UploadFile = File(...), header_row: int | None = Form(None)):
    try:
        raw = await file.read()
        return template_store.detect_columns(raw, header_row)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Fichier illisible : {e}")


@app.post("/comments")
async def save_comment_route(payload: dict):
    code = str(payload.get("code", "")).strip()
    lot = str(payload.get("lot", "")).strip()
    if not code or not lot:
        raise HTTPException(status_code=400, detail="code et lot sont obligatoires.")
    text = payload.get("text", "")
    updated = str(payload.get("inventory_date", ""))
    comment_store.set_comment(code, lot, text, updated)
    return {"ok": True}


@app.get("/health")
async def health():
    return {"status": "ok"}
